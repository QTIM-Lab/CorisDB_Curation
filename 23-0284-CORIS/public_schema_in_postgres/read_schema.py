import psycopg2
import pandas as pd
import os
from openpyxl import load_workbook
import re

# Database connection parameters
db_params = {
    'dbname': 'coris_db',
    'user': 'coris_admin',
    'password': '',
    'host': 'localhost',
    'port': '5432'
}

# Output Excel file path
excel_file_path = '/scratch90/QTIM/Active/23-0284/EHR/CorisDB_Curation/23-0284-CORIS/public_schema_in_postgres/coris_schema_type.xlsx'

# Ensure the output directory exists
os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)

# Natural sort key function
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

try:
    # Connect to the PostgreSQL database
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    # Get all table names in the coris_registry schema ending with '20250429'
    cur.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'coris_registry' AND table_name LIKE '%20250429'
    """)
    tables = cur.fetchall()

    # Create an Excel writer
    with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
        for table in tables:
            table_name = table[0]
            cur.execute("""
                SELECT column_name, data_type
                FROM information_schema.columns
                WHERE table_schema = 'coris_registry' AND table_name = %s
            """, (table_name,))
            columns = cur.fetchall()

            # Create DataFrame and write to a sheet
            df = pd.DataFrame(columns, columns=['Column Name', 'Data Type'])
            df.to_excel(writer, sheet_name=table_name[:31], index=False)

    # Reorder sheets using natural sort
    wb = load_workbook(excel_file_path)
    sorted_sheet_names = sorted(wb.sheetnames, key=natural_sort_key)
    wb._sheets = [wb[sheet] for sheet in sorted_sheet_names]
    wb.save(excel_file_path)

    print(f"✅ Schema information saved and sheets reordered naturally at: {excel_file_path}")

except Exception as e:
    print(f"❌ Error: {e}")

finally:
    if 'cur' in locals():
        cur.close()
    if 'conn' in locals():
        conn.close()
